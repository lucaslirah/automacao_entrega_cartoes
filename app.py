import io
import string
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

#Nome do repositorio : automacao-entrega-cartoes

# ---------------------------------------------------------
# CONFIGURAÇÃO DA PÁGINA STREAMLIT
# ---------------------------------------------------------
st.set_page_config(
    page_title="Automação de Entrega de Cartões",
    page_icon="💳",
    layout="wide"
)

# ---------------------------------------------------------
# FUNÇÃO PRINCIPAL DE PROCESSAMENTO DOS DADOS
# ---------------------------------------------------------
def processar_planilha(arquivo_carregado, limite_dia=180, total_guiches=6):
    """
    Processa o arquivo Excel carregado, aplica o algoritmo de dias/guichês
    e gera a planilha final formatada em memória.
    """
    # 1. Leitura da aba DADOS
    df_dados = pd.read_excel(arquivo_carregado, sheet_name="DADOS")
    
    # Normalização dos cabeçalhos
    df_dados.columns = [str(col).strip() for col in df_dados.columns]

    # Identificação flexível da coluna de nomes
    coluna_nome = None
    for col in df_dados.columns:
        if col.upper() in ["NOME", "NOME COMPLETO", "NOME_COMPLETO", "CLIENTE", "TITULAR"]:
            coluna_nome = col
            break

    if not coluna_nome:
        raise ValueError("A coluna de NOME não foi encontrada na aba DADOS.")

    df_dados = df_dados.rename(columns={coluna_nome: "NOME"})

    # Identificação flexível de outras colunas
    for col in df_dados.columns:
        col_upper = col.upper()
        if "CPF" in col_upper and col != "CPF":
            df_dados = df_dados.rename(columns={col: "CPF"})
        elif "CONTA" in col_upper and col != "CONTA":
            df_dados = df_dados.rename(columns={col: "CONTA"})
        elif "ABREVIADO" in col_upper and col != "NOME_ABREVIADO":
            df_dados = df_dados.rename(columns={col: "NOME_ABREVIADO"})

    for col_opcional in ["CPF", "CONTA", "NOME_ABREVIADO"]:
        if col_opcional not in df_dados.columns:
            df_dados[col_opcional] = ""

    # Limpeza e remoção de registros vazios
    df_dados["NOME"] = df_dados["NOME"].astype(str).str.strip()
    df_dados = df_dados[df_dados["NOME"] != ""].copy()

    # 2. Ordenação A -> Z e definição da ORDEM original
    df_dados = df_dados.sort_values(by="NOME").reset_index(drop=True)
    df_dados["ORDEM"] = range(1, len(df_dados) + 1)
    df_dados["INICIAL"] = df_dados["NOME"].str[0].str.upper()

    # 3. Algoritmo: Agrupamento por Dia
    contagem_series = df_dados["INICIAL"].value_counts()
    alfabeto = list(string.ascii_uppercase)
    dados_letras = [{"letra": l, "quantidade": int(contagem_series.get(l, 0))} for l in alfabeto]

    dias_agrupados = []
    dia_atual_letras = []
    acumulado_dia = 0
    mapa_letra_dia = {}

    for item in dados_letras:
        letra = item["letra"]
        qtd = item["quantidade"]

        if (acumulado_dia + qtd > limite_dia) and len(dia_atual_letras) > 0:
            dias_agrupados.append(dia_atual_letras)
            dia_atual_letras = []
            acumulado_dia = 0

        dia_atual_letras.append(item)
        acumulado_dia += qtd

    if dia_atual_letras:
        dias_agrupados.append(dia_atual_letras)

    for num_dia, grupo in enumerate(dias_agrupados, start=1):
        for item in grupo:
            mapa_letra_dia[item["letra"]] = num_dia

    df_dados["DIA"] = df_dados["INICIAL"].map(mapa_letra_dia)

    # 4. Algoritmo: Divisão em Guichês
    df_dados["GUICHE"] = 0
    for dia in sorted(df_dados["DIA"].unique()):
        idx_dia = df_dados[df_dados["DIA"] == dia].index
        total_dia = len(idx_dia)
        tamanho_bloco = total_dia // total_guiches

        for g in range(1, total_guiches + 1):
            inicio = (g - 1) * tamanho_bloco
            fim = g * tamanho_bloco if g < total_guiches else total_dia
            indices_guiche = idx_dia[inicio:fim]
            df_dados.loc[indices_guiche, "GUICHE"] = g

    # 5. Geração do arquivo Excel com formatação OpenpyXL
    wb = openpyxl.load_workbook(arquivo_carregado)

    fonte_cabecalho = Font(name="Calibri", size=11, bold=True, color="000000")
    fonte_corpo = Font(name="Calibri", size=11)
    fonte_total = Font(name="Calibri", size=11, bold=True)
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    preenchimento_cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    borda_fina = Side(border_style="thin", color="000000")
    borda_completa = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    # --- ABA LETRAS_DIAS ---
    if "LETRAS_DIAS" in wb.sheetnames:
        del wb["LETRAS_DIAS"]
    ws_letras = wb.create_sheet(title="LETRAS_DIAS", index=0)

    ws_letras.merge_cells("B1:E1")
    ws_letras["B1"] = "CONTAS NOVAS"
    ws_letras["B1"].font = fonte_cabecalho
    ws_letras["B1"].alignment = alinhamento_centro
    ws_letras["B1"].fill = preenchimento_cinza
    for col in range(2, 6):
        ws_letras.cell(row=1, column=col).border = borda_completa

    titulos_letras = ["DIA", "INICIAIS", "UNITARIO", "TOTAL"]
    for i, t in enumerate(titulos_letras, start=2):
        c = ws_letras.cell(row=2, column=i, value=t)
        c.font = fonte_cabecalho
        c.alignment = alinhamento_centro
        c.fill = preenchimento_cinza
        c.border = borda_completa

    linha_atual = 3
    total_geral = 0
    for num_dia, grupo in enumerate(dias_agrupados, start=1):
        linha_ini = linha_atual
        soma_dia = sum(it["quantidade"] for it in grupo)
        total_geral += soma_dia

        for it in grupo:
            c3 = ws_letras.cell(row=linha_atual, column=3, value=it["letra"])
            c4 = ws_letras.cell(row=linha_atual, column=4, value=it["quantidade"])
            c3.alignment = c4.alignment = alinhamento_centro
            c3.font = c4.font = fonte_corpo
            c3.border = c4.border = borda_completa
            linha_atual += 1

        linha_fim = linha_atual - 1
        ws_letras.merge_cells(start_row=linha_ini, start_column=2, end_row=linha_fim, end_column=2)
        cd = ws_letras.cell(row=linha_ini, column=2, value=num_dia)
        cd.alignment = alinhamento_centro
        cd.font = fonte_cabecalho

        ws_letras.merge_cells(start_row=linha_ini, start_column=5, end_row=linha_fim, end_column=5)
        ct = ws_letras.cell(row=linha_ini, column=5, value=soma_dia)
        ct.alignment = alinhamento_centro
        ct.font = fonte_cabecalho

        for r in range(linha_ini, linha_fim + 1):
            ws_letras.cell(row=r, column=2).border = borda_completa
            ws_letras.cell(row=r, column=5).border = borda_completa

    ws_letras.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=4)
    c_rodape = ws_letras.cell(row=linha_atual, column=2, value="TOTAL GERAL")
    c_rodape.font = fonte_total
    c_rodape.alignment = alinhamento_centro
    c_rodape.fill = preenchimento_cinza

    c_tot = ws_letras.cell(row=linha_atual, column=5, value=total_geral)
    c_tot.font = fonte_total
    c_tot.alignment = alinhamento_centro
    c_tot.fill = preenchimento_cinza

    for col in range(2, 6):
        ws_letras.cell(row=linha_atual, column=col).border = borda_completa

    ws_letras.column_dimensions["A"].width = 3
    ws_letras.column_dimensions["B"].width = 12
    ws_letras.column_dimensions["C"].width = 12
    ws_letras.column_dimensions["D"].width = 14
    ws_letras.column_dimensions["E"].width = 14

    # --- ABA DADOS (Pilha Reversa) ---
    if "DADOS" in wb.sheetnames:
        del wb["DADOS"]
    ws_dados = wb.create_sheet(title="DADOS", index=1)

    colunas_dados_finais = ["ORDEM", "CONTA", "NOME", "NOME_ABREVIADO", "CPF", "DIA"]
    for c_idx, col_nome in enumerate(colunas_dados_finais, start=1):
        cel = ws_dados.cell(row=1, column=c_idx, value=col_nome)
        cel.font = fonte_cabecalho
        cel.alignment = alinhamento_centro
        cel.fill = preenchimento_cinza
        cel.border = borda_completa

    df_dados_invertido = df_dados.iloc[::-1].reset_index(drop=True)
    for r_idx, row in df_dados_invertido.iterrows():
        linha_num = r_idx + 2
        valores = [
            row["ORDEM"],
            str(row.get("CONTA", "")),
            row["NOME"],
            str(row.get("NOME_ABREVIADO", "")),
            str(row.get("CPF", "")),
            row["DIA"]
        ]
        for c_idx, val in enumerate(valores, start=1):
            cel = ws_dados.cell(row=linha_num, column=c_idx, value=val)
            cel.font = fonte_corpo
            cel.border = borda_completa
            if c_idx in [3, 4]:
                cel.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cel.alignment = alinhamento_centro

    ws_dados.column_dimensions["A"].width = 10
    ws_dados.column_dimensions["B"].width = 16
    ws_dados.column_dimensions["C"].width = 38
    ws_dados.column_dimensions["D"].width = 28
    ws_dados.column_dimensions["E"].width = 18
    ws_dados.column_dimensions["F"].width = 8

    # --- ABAS POR GUICHÊ (Pilha Reversa) ---
    for g in range(1, total_guiches + 1):
        nome_aba = f"GUICHE_{g}"
        if nome_aba in wb.sheetnames:
            del wb[nome_aba]
        
        ws_g = wb.create_sheet(title=nome_aba)
        df_guiche = df_dados[df_dados["GUICHE"] == g].iloc[::-1].reset_index(drop=True)

        colunas_guiche = ["ORDEM", "DIA", "NOME", "CPF", "CONTA"]
        for c_idx, col_nome in enumerate(colunas_guiche, start=1):
            cel = ws_g.cell(row=1, column=c_idx, value=col_nome)
            cel.font = fonte_cabecalho
            cel.alignment = alinhamento_centro
            cel.fill = preenchimento_cinza
            cel.border = borda_completa

        for r_idx, row in df_guiche.iterrows():
            linha_num = r_idx + 2
            valores = [
                row["ORDEM"],
                row["DIA"],
                row["NOME"],
                str(row.get("CPF", "")),
                str(row.get("CONTA", ""))
            ]
            for c_idx, val in enumerate(valores, start=1):
                cel = ws_g.cell(row=linha_num, column=c_idx, value=val)
                cel.font = fonte_corpo
                cel.border = borda_completa
                if c_idx == 3:
                    cel.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cel.alignment = alinhamento_centro

        ws_g.column_dimensions["A"].width = 10
        ws_g.column_dimensions["B"].width = 8
        ws_g.column_dimensions["C"].width = 38
        ws_g.column_dimensions["D"].width = 18
        ws_g.column_dimensions["E"].width = 18

    # Salva o resultado em buffer binário
    buffer_saida = io.BytesIO()
    wb.save(buffer_saida)
    buffer_saida.seek(0)
    
    return buffer_saida, df_dados, dias_agrupados

# ---------------------------------------------------------
# CONSTRUÇÃO DA INTERFACE WEB
# ---------------------------------------------------------
st.title("Automação de Entrega de Cartões")
st.markdown("Faça o upload da planilha base com a aba **DADOS** para gerar a separação por **Dias** e **Guichês** com ordenação em pilha reversa.")

# Barra lateral para parâmetros
with st.sidebar:
    st.header("Parâmetros de Configuração")
    limite_input = st.number_input("Limite máximo de atendimentos/dia:", min_value=50, max_value=500, value=180, step=10)
    guiches_input = st.number_input("Quantidade de Guichês:", min_value=1, max_value=20, value=6, step=1)
    st.info("💡 As configurações padrão utilizam 180 cartões/dia e 6 guichês.")

# Upload do arquivo Excel
arquivo = st.file_uploader("Selecione ou arraste a planilha Excel (.xlsx)", type=["xlsx"])

if arquivo is not None:
    st.success(f"Arquivo carregado com sucesso: **{arquivo.name}**")
    
    with st.spinner("Processando dados e aplicando algoritmo de distribuição..."):
        try:
            excel_gerado, df_processado, grupos_dias = processar_planilha(
                arquivo, 
                limite_dia=limite_input, 
                total_guiches=guiches_input
            )
            
            total_contas = len(df_processado)
            total_dias = len(grupos_dias)
            
            # Métricas no topo da página
            col1, col2, col3 = st.columns(3)
            col1.metric("Total de Contas", f"{total_contas:,}".replace(",", "."))
            col2.metric("Total de Dias Úteis", total_dias)
            col3.metric("Média por Dia", f"{total_contas / total_dias:.1f}")

            # Botão de Download em Destaque
            st.download_button(
                label="📥 Baixar Planilha Processada (.xlsx)",
                data=excel_gerado,
                file_name=f"PROCESSADA_{arquivo.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

            # Resumos e visualizações
            st.divider()
            tab1, tab2 = st.tabs(["📅 Distribuição por Dia", "🏪 Distribuição por Guichê"])
            
            with tab1:
                resumo_dias = df_processado.groupby("DIA")["NOME"].count().reset_index()
                resumo_dias.columns = ["Dia", "Total de Contas"]
                st.dataframe(resumo_dias, use_container_width=True)

            with tab2:
                crosstab_guiches = pd.crosstab(df_processado["DIA"], df_processado["GUICHE"], margins=True, margins_name="Total")
                st.dataframe(crosstab_guiches, use_container_width=True)

        except Exception as e:
            st.error(f"Erro ao processar o arquivo: {e}")