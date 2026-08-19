import os
import glob
import pandas as pd
import string
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("\n" + "="*60)
print("   AUTOMAÇÃO: SEPARAÇÃO DE CARTÕES COM PILHA REVERSA")
print("="*60)

# ---------------------------------------------------------
# 1. LOCALIZAÇÃO AUTOMÁTICA DA PLANILHA EXCEL
# ---------------------------------------------------------
arquivos_excel = [
    f for f in glob.glob("*.xlsx") 
    if not os.path.basename(f).startswith("~$")
]

if not arquivos_excel:
    print("[ERRO] Nenhum arquivo .xlsx encontrado na pasta atual!")
    exit()

caminho_arquivo = arquivos_excel[0]
print(f"[INFO] Arquivo selecionado: '{caminho_arquivo}'")

# ---------------------------------------------------------
# 2. LEITURA E PADRONIZAÇÃO DAS COLUNAS
# ---------------------------------------------------------
try:
    df_dados = pd.read_excel(caminho_arquivo, sheet_name="DADOS")
    print(f"[INFO] Aba 'DADOS' carregada com {len(df_dados)} registros.")
except Exception as e:
    print(f"[ERRO] Falha ao ler a aba 'DADOS': {e}")
    exit()

# Normaliza os nomes de colunas
df_dados.columns = [str(col).strip() for col in df_dados.columns]

coluna_nome = None
for col in df_dados.columns:
    if col.upper() in ["NOME", "NOME COMPLETO", "NOME_COMPLETO", "CLIENTE", "TITULAR"]:
        coluna_nome = col
        break

if not coluna_nome:
    print(f"[ERRO] Coluna de nomes não encontrada. Colunas lidas: {list(df_dados.columns)}")
    exit()

df_dados = df_dados.rename(columns={coluna_nome: "NOME"})

# Localiza e padroniza CPF, CONTA e NOME_ABREVIADO
for col in df_dados.columns:
    col_upper = col.upper()
    if "CPF" in col_upper and col != "CPF":
        df_dados = df_dados.rename(columns={col: "CPF"})
    elif "CONTA" in col_upper and col != "CONTA":
        df_dados = df_dados.rename(columns={col: "CONTA"})
    elif "ABREVIADO" in col_upper and col != "NOME_ABREVIADO":
        df_dados = df_dados.rename(columns={col: "NOME_ABREVIADO"})

for col_opcional in ["CPF", "CONTA", "NOME_ABREVIADO"]:
    if col_opcional not in df_dados.columns:
        df_dados[col_opcional] = ""

# Limpeza e remoção de registros vazios
df_dados["NOME"] = df_dados["NOME"].astype(str).str.strip()
df_dados = df_dados[df_dados["NOME"] != ""].copy()

# ---------------------------------------------------------
# 3. ORDENAÇÃO ALFABÉTICA (A -> Z) E DEFINIÇÃO DA ORDEM
# ---------------------------------------------------------
df_dados = df_dados.sort_values(by="NOME").reset_index(drop=True)
df_dados["ORDEM"] = range(1, len(df_dados) + 1)
df_dados["INICIAL"] = df_dados["NOME"].str[0].str.upper()

# ---------------------------------------------------------
# 4. ALGORITMO: AGRUPAMENTO POR DIA (LIMITE <= 200)
# ---------------------------------------------------------
LIMITE_MAXIMO_DIA = 200
contagem_series = df_dados["INICIAL"].value_counts()
alfabeto = list(string.ascii_uppercase)

dados_letras = [{"letra": l, "quantidade": int(contagem_series.get(l, 0))} for l in alfabeto]

dias_agrupados = []
dia_atual_letras = []
acumulado_dia = 0
mapa_letra_dia = {}

for item in dados_letras:
    letra = item["letra"]
    qtd = item["quantidade"]

    if (acumulado_dia + qtd > LIMITE_MAXIMO_DIA) and len(dia_atual_letras) > 0:
        dias_agrupados.append(dia_atual_letras)
        dia_atual_letras = []
        acumulado_dia = 0

    dia_atual_letras.append(item)
    acumulado_dia += qtd

if dia_atual_letras:
    dias_agrupados.append(dia_atual_letras)

for num_dia, grupo in enumerate(dias_agrupados, start=1):
    for item in grupo:
        mapa_letra_dia[item["letra"]] = num_dia

df_dados["DIA"] = df_dados["INICIAL"].map(mapa_letra_dia)

# ---------------------------------------------------------
# 5. ALGORITMO: DIVISÃO EQUITATIVA EM 6 GUICHÊS
# ---------------------------------------------------------
TOTAL_GUICHES = 6
df_dados["GUICHE"] = 0

for dia in sorted(df_dados["DIA"].unique()):
    idx_dia = df_dados[df_dados["DIA"] == dia].index
    total_dia = len(idx_dia)
    tamanho_bloco = total_dia // TOTAL_GUICHES

    for g in range(1, TOTAL_GUICHES + 1):
        inicio = (g - 1) * tamanho_bloco
        fim = g * tamanho_bloco if g < TOTAL_GUICHES else total_dia
        indices_guiche = idx_dia[inicio:fim]
        df_dados.loc[indices_guiche, "GUICHE"] = g

print(f"[INFO] Distribuição concluída: {len(df_dados)} contas em {len(dias_agrupados)} dias e {TOTAL_GUICHES} guichês.")

# ---------------------------------------------------------
# 6. MONTAGEM E FORMATAÇÃO NO EXCEL COM OPENPYXL
# ---------------------------------------------------------
wb = openpyxl.load_workbook(caminho_arquivo)

# Estilos Visuais
fonte_cabecalho = Font(name="Calibri", size=11, bold=True, color="000000")
fonte_corpo = Font(name="Calibri", size=11)
fonte_total = Font(name="Calibri", size=11, bold=True)
alinhamento_centro = Alignment(horizontal="center", vertical="center")
preenchimento_cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
borda_fina = Side(border_style="thin", color="000000")
borda_completa = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

# A) ABA LETRAS_DIAS
if "LETRAS_DIAS" in wb.sheetnames:
    del wb["LETRAS_DIAS"]
ws_letras = wb.create_sheet(title="LETRAS_DIAS", index=0)

# Título Superior
ws_letras.merge_cells("B1:E1")
ws_letras["B1"] = "CONTAS NOVAS"
ws_letras["B1"].font = fonte_cabecalho
ws_letras["B1"].alignment = alinhamento_centro
ws_letras["B1"].fill = preenchimento_cinza
for col in range(2, 6):
    ws_letras.cell(row=1, column=col).border = borda_completa

# Cabeçalhos
titulos_letras = ["DIA", "INICIAIS", "UNITARIO", "TOTAL"]
for i, t in enumerate(titulos_letras, start=2):
    c = ws_letras.cell(row=2, column=i, value=t)
    c.font = fonte_cabecalho
    c.alignment = alinhamento_centro
    c.fill = preenchimento_cinza
    c.border = borda_completa

linha_atual = 3
total_geral = 0
for num_dia, grupo in enumerate(dias_agrupados, start=1):
    linha_ini = linha_atual
    soma_dia = sum(it["quantidade"] for it in grupo)
    total_geral += soma_dia

    for it in grupo:
        c3 = ws_letras.cell(row=linha_atual, column=3, value=it["letra"])
        c4 = ws_letras.cell(row=linha_atual, column=4, value=it["quantidade"])
        c3.alignment = c4.alignment = alinhamento_centro
        c3.font = c4.font = fonte_corpo
        c3.border = c4.border = borda_completa
        linha_atual += 1

    linha_fim = linha_atual - 1
    
    # Mescla DIA e TOTAL
    ws_letras.merge_cells(start_row=linha_ini, start_column=2, end_row=linha_fim, end_column=2)
    cd = ws_letras.cell(row=linha_ini, column=2, value=num_dia)
    cd.alignment = alinhamento_centro
    cd.font = fonte_cabecalho

    ws_letras.merge_cells(start_row=linha_ini, start_column=5, end_row=linha_fim, end_column=5)
    ct = ws_letras.cell(row=linha_ini, column=5, value=soma_dia)
    ct.alignment = alinhamento_centro
    ct.font = fonte_cabecalho

    for r in range(linha_ini, linha_fim + 1):
        ws_letras.cell(row=r, column=2).border = borda_completa
        ws_letras.cell(row=r, column=5).border = borda_completa

# Rodapé TOTAL GERAL
ws_letras.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=4)
c_rodape = ws_letras.cell(row=linha_atual, column=2, value="TOTAL GERAL")
c_rodape.font = fonte_total
c_rodape.alignment = alinhamento_centro
c_rodape.fill = preenchimento_cinza

c_tot = ws_letras.cell(row=linha_atual, column=5, value=total_geral)
c_tot.font = fonte_total
c_tot.alignment = alinhamento_centro
c_tot.fill = preenchimento_cinza

for col in range(2, 6):
    ws_letras.cell(row=linha_atual, column=col).border = borda_completa

ws_letras.column_dimensions["A"].width = 3
ws_letras.column_dimensions["B"].width = 12
ws_letras.column_dimensions["C"].width = 12
ws_letras.column_dimensions["D"].width = 14
ws_letras.column_dimensions["E"].width = 14

# B) ABA DADOS (Invertida: do último 906 ao primeiro 1)
if "DADOS" in wb.sheetnames:
    del wb["DADOS"]
ws_dados = wb.create_sheet(title="DADOS", index=1)

colunas_dados_finais = ["ORDEM", "CONTA", "NOME", "NOME_ABREVIADO", "CPF", "DIA"]

for c_idx, col_nome in enumerate(colunas_dados_finais, start=1):
    cel = ws_dados.cell(row=1, column=c_idx, value=col_nome)
    cel.font = fonte_cabecalho
    cel.alignment = alinhamento_centro
    cel.fill = preenchimento_cinza
    cel.border = borda_completa

# Inversão da aba DADOS (Pilha Reversa)
df_dados_invertido = df_dados.iloc[::-1].reset_index(drop=True)

for r_idx, row in df_dados_invertido.iterrows():
    linha_num = r_idx + 2
    valores = [
        row["ORDEM"],
        str(row.get("CONTA", "")),
        row["NOME"],
        str(row.get("NOME_ABREVIADO", "")),
        str(row.get("CPF", "")),
        row["DIA"]
    ]
    for c_idx, val in enumerate(valores, start=1):
        cel = ws_dados.cell(row=linha_num, column=c_idx, value=val)
        cel.font = fonte_corpo
        cel.border = borda_completa
        if c_idx in [3, 4]:
            cel.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cel.alignment = alinhamento_centro

ws_dados.column_dimensions["A"].width = 10
ws_dados.column_dimensions["B"].width = 16
ws_dados.column_dimensions["C"].width = 38
ws_dados.column_dimensions["D"].width = 28
ws_dados.column_dimensions["E"].width = 18
ws_dados.column_dimensions["F"].width = 8

# C) ABAS POR GUICHÊ (GUICHE_1 até GUICHE_6) - Invertidas por Guichê
for g in range(1, TOTAL_GUICHES + 1):
    nome_aba = f"GUICHE_{g}"
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]
    
    ws_g = wb.create_sheet(title=nome_aba)
    
    # Filtra as contas do guichê e inverte a ordem (pilha reversa)
    df_guiche = df_dados[df_dados["GUICHE"] == g].iloc[::-1].reset_index(drop=True)

    colunas_guiche = ["ORDEM", "DIA", "NOME", "CPF", "CONTA"]
    for c_idx, col_nome in enumerate(colunas_guiche, start=1):
        cel = ws_g.cell(row=1, column=c_idx, value=col_nome)
        cel.font = fonte_cabecalho
        cel.alignment = alinhamento_centro
        cel.fill = preenchimento_cinza
        cel.border = borda_completa

    for r_idx, row in df_guiche.iterrows():
        linha_num = r_idx + 2
        valores = [
            row["ORDEM"],
            row["DIA"],
            row["NOME"],
            str(row.get("CPF", "")),
            str(row.get("CONTA", ""))
        ]
        for c_idx, val in enumerate(valores, start=1):
            cel = ws_g.cell(row=linha_num, column=c_idx, value=val)
            cel.font = fonte_corpo
            cel.border = borda_completa
            if c_idx == 3:
                cel.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cel.alignment = alinhamento_centro

    ws_g.column_dimensions["A"].width = 10
    ws_g.column_dimensions["B"].width = 8
    ws_g.column_dimensions["C"].width = 38
    ws_g.column_dimensions["D"].width = 18
    ws_g.column_dimensions["E"].width = 18

# Salva arquivo final
wb.save(caminho_arquivo)
print(f"[SUCESSO] Planilha '{os.path.basename(caminho_arquivo)}' processada no formato de pilha reversa (906 -> 1)!")
print("="*60 + "\n")